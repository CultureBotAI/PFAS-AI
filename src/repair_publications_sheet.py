#!/usr/bin/env python3
"""Repair publications sheet in PFAS Data for AI.xlsx

This script fixes the duplicate header issue in the publications sheet by:
1. Creating a unified schema with all columns from both sections
2. Merging manual and auto-generated publications into a single table
3. Adding proper source provenance tracking
"""

import pandas as pd
import openpyxl
from pathlib import Path
import shutil
from datetime import datetime

def repair_publications_sheet(excel_path: str, backup: bool = True) -> None:
    """Repair the publications sheet with unified schema.

    Args:
        excel_path: Path to the Excel file
        backup: Whether to create a backup before modifying (default: True)
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Create backup
    if backup:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = excel_path.parent / f"{excel_path.stem}_backup_{timestamp}{excel_path.suffix}"
        shutil.copy2(excel_path, backup_path)
        print(f"✓ Created backup: {backup_path}")

    # Read the Excel file
    print(f"\nReading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    # Check if publications sheet exists
    sheet_name = "publications"
    if sheet_name not in wb.sheetnames:
        print(f"✗ Sheet '{sheet_name}' not found in workbook")
        print(f"  Available sheets: {wb.sheetnames}")
        return

    ws = wb[sheet_name]

    # Read all data from sheet
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    print(f"✓ Read {len(data)} rows from publications sheet")

    # Analyze structure
    header1 = data[0] if len(data) > 0 else []
    print(f"\nOriginal structure:")
    print(f"  Row 0 (Header 1): {header1}")
    print(f"  Total rows: {len(data)}")

    # Find duplicate header
    duplicate_header_row = None
    for i, row in enumerate(data):
        if i == 0:
            continue
        # Check if row looks like a header (has 'url' or 'URL' in first cell)
        if row and row[0] and str(row[0]).lower() in ['url', 'pmid', 'title']:
            # Additional check: next row after this should have data or be empty
            if i < len(data) - 1:
                duplicate_header_row = i
                print(f"  Row {i} (Duplicate header): {row}")
                break

    # Define unified schema
    unified_header = ['url', 'title', 'journal', 'year', 'authors', 'pdf', 'search_term', 'pmid', 'source']
    print(f"\nUnified header ({len(unified_header)} columns):")
    print(f"  {unified_header}")

    # Split data into sections
    if duplicate_header_row:
        manual_data = data[1:duplicate_header_row]  # After header1, before duplicate
        # Skip empty rows around duplicate header
        auto_start = duplicate_header_row + 1
        while auto_start < len(data) and not any(data[auto_start]):
            auto_start += 1
        auto_data = data[auto_start:] if auto_start < len(data) else []

        # Filter out empty rows from both sections
        manual_data = [row for row in manual_data if any(row)]
        auto_data = [row for row in auto_data if any(row)]

        print(f"\nData sections:")
        print(f"  Manual publications: {len(manual_data)} rows")
        print(f"  Auto-generated publications: {len(auto_data)} rows")
    else:
        # No duplicate header found - just clean up the existing data
        manual_data = [row for row in data[1:] if any(row)]
        auto_data = []
        print(f"\nNo duplicate header found")
        print(f"  All publications: {len(manual_data)} rows")

    # Transform manual data to unified schema
    unified_data = []

    for row in manual_data:
        # Extend row to have at least 6 columns
        while len(row) < 6:
            row.append(None)

        unified_row = [
            row[0],  # url
            row[1],  # title
            row[2] if len(row) > 2 else None,  # journal
            row[3] if len(row) > 3 else None,  # year
            row[4] if len(row) > 4 else None,  # authors
            row[5] if len(row) > 5 else None,  # pdf
            None,  # search_term (empty for manual)
            None,  # pmid (empty for manual)
            'manual'  # source
        ]
        unified_data.append(unified_row)

    # Transform auto data to unified schema
    for row in auto_data:
        # Extend row to have at least 5 columns
        while len(row) < 5:
            row.append(None)

        unified_row = [
            row[0],  # url
            row[1],  # title
            None,  # journal (empty for auto)
            None,  # year (empty for auto - might be in title)
            None,  # authors (empty for auto)
            None,  # pdf (empty for auto)
            row[2] if len(row) > 2 else None,  # search_term
            row[3] if len(row) > 3 else None,  # pmid
            row[4] if len(row) > 4 else 'extend1'  # source
        ]
        unified_data.append(unified_row)

    print(f"\nUnified data:")
    print(f"  Total rows: {len(unified_data)}")

    # Clear the sheet
    ws.delete_rows(1, ws.max_row)

    # Write unified header
    for col_idx, header_val in enumerate(unified_header, start=1):
        ws.cell(row=1, column=col_idx, value=header_val)

    # Write unified data
    for row_idx, row_data in enumerate(unified_data, start=2):
        for col_idx, cell_val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_val)

    # Save the workbook
    wb.save(excel_path)
    print(f"\n✓ Saved repaired Excel file: {excel_path}")
    print(f"  Final structure: {len(unified_data)} data rows + 1 header row")
    print(f"  Columns: {len(unified_header)}")

    # Verify by reading back
    df = pd.read_excel(excel_path, sheet_name='publications')
    print(f"\nVerification:")
    print(f"  DataFrame shape: {df.shape}")
    print(f"  Columns: {list(df.columns)}")
    print(f"  Source distribution:")
    if 'source' in df.columns:
        print(df['source'].value_counts().to_string(header=False).replace('\n', '\n    '))

    print(f"\n✓ Publications sheet repair complete!")


if __name__ == "__main__":
    import sys

    # Default path
    excel_path = "data/sheet/PFAS Data for AI.xlsx"

    # Allow override from command line
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]

    print("=" * 70)
    print("PFAS Data for AI - Publications Sheet Repair")
    print("=" * 70)

    try:
        repair_publications_sheet(excel_path, backup=True)
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
